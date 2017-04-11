#!/usr/bin/python
from sys import argv
import openpyxl
import sys, getopt

def main(argv):
	print "called"
	iplist_file = ''
	workbook = ''
	column_name = ''
	try:
		opts,args = getopt.getopt(argv,"hf:w:c:",["iplist_file=","workbook=","column_name="])
		total = len(opts)
		print "total no of arg %d" %total
	except getopt.GetError:
		print "Error"
	for opt, arg in opts:
		if opt == '-f':
			iplist_file = arg
			print "iplist file is %s" %iplist_file
		elif opt == '-w':
			workbook = arg
			print "workbook name is %s" %workbook
		elif opt == '-c':
			column_name = arg
			print "column name is %s" %column_name

	wb = openpyxl.load_workbook(workbook)
	sheet = wb.get_sheet_by_name('Sheet1')
	f = open(iplist_file,'r')
	for line in f:
		for i in range(sheet.min_column,sheet.max_column+1):
			x = sheet.cell(row = 1, column = i).value
			if x == column_name:
				print "Found column name match"
				for j in range(sheet.min_row+1,sheet.max_row+1):
					fruit_name = str(sheet.cell(row = j, column = i).value)
					print "line is %s" %line
					print "fruit name is %s" %fruit_name
					print type(fruit_name)
					print type(line)
					print fruit_name == line
				
				#print sheet.cell(row=1,column = i-1).value

if __name__ == '__main__':
	main(sys.argv[1:])
